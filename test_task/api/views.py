from django.http import HttpResponse
from django.shortcuts import render, redirect
from openpyxl.styles import (Font, Alignment, NamedStyle,
                             PatternFill, Border, Side)
from rest_framework import views
import openpyxl
import pandas as pd


side_config = Side(border_style="thin", color='A0A0A0')
header = NamedStyle(name="header")
header.font = Font(name='Arial', color="00005C", size=10, bold=True)
header.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
header.fill = PatternFill(fill_type='solid', fgColor='CBE4E5')
header.border = Border(top=side_config, right=side_config, bottom=side_config, left=side_config)


def index(request):
    """View для главной страницы."""

    return render(request, 'index.html')



def calculating_tax(tax_base):
    """Исчисление налога."""

    if tax_base > 0:
        if tax_base <= 5000000:
            return round(int(tax_base*100)*0.13/100, 0)
        return round(int(tax_base*100)*0.15/100, 0)
    return None


def calculating_deviation(input_tax, calculate_tax):
    """Вычисление отклонения."""

    if isinstance(input_tax, int) or isinstance(input_tax, float):
        if input_tax > 0:
            return (int(round((int(input_tax*100) 
                    - int(calculate_tax*100))/100, 0)))
    return None


def create_sheet(output_data):
    """Создание шаблона листа для отчета."""

    ws = output_data.active
    ws.title = "Отчет"
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 34
    ws.column_dimensions['C'].width = 13
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 17
    ws.column_dimensions['F'].width = 13
    ws.row_dimensions[1].height = 13
    ws.row_dimensions[2].height = 26
    ws["A1"] = 'Филиал'
    ws.merge_cells('A1:A2')
    ws["B1"] = 'Сотрудник'
    ws.merge_cells('B1:B2')
    ws["C1"] = 'Налоговая база'
    ws.merge_cells('C1:C2')
    ws["D1"] = 'Налог'
    ws.merge_cells('D1:E1')
    ws["F1"] = 'Отклонение'
    ws.merge_cells('F1:F2')
    ws["D2"] = 'Исчислено всего'
    ws["E2"] = 'Исчислено всего по формуле всего'
    columns = ['A', 'B', 'C', 'D', 'E', 'F']
    for i in range(1,3):
        for l in columns:
            ws[f'{l}{i}'].style = header
    return ws


def sort_column(name_file):
    """Сортировка по столбцу определенному в column_name."""

    column_name = 'Отклонение'
    data = pd.DataFrame(pd.read_excel(name_file))
    sort_data = data.sort_values(column_name, ascending=False)
    sort_data.to_excel('done.xlsx')


class СheckView(views.APIView):
    """View для обработки получаемых данных."""

    def post(self, request):
        try:
            input_data = pd.read_excel(request.FILES['input_data'])
            input_data_dict = input_data.to_dict(orient='records')
            output_data = openpyxl.Workbook()
            ws = create_sheet(output_data)
            for i in range(2, len(input_data_dict)):
                row = input_data_dict[i]
                tax = calculating_tax(row['Налоговая база'])
                deviation = calculating_deviation(row['Налог'], tax)
                ws[f"A{i+1}"] = row['Филиал']
                ws[f"B{i+1}"] = row['Сотрудник']
                ws[f"C{i+1}"] = row['Налоговая база']
                ws[f"D{i+1}"] = row['Налог']
                ws[f"E{i+1}"] = tax
                ws[f"F{i+1}"] = deviation
            ws_sort = output_data['Отчет']
            ws_sort.auto_filter.ref = f"F1:F{ws.max_row}"
            ws_sort.auto_filter.add_sort_condition(f"F1:F{ws.max_row}")
            for i in range(3, ws.max_row+1):
                if ws[f'F{i}'].value == 0 or ws[f'F{i}'].value == None:
                    ws[f'F{i}'].fill = PatternFill(fill_type='solid', fgColor='00B050')
                else:
                    ws[f'F{i}'].fill = PatternFill(fill_type='solid', fgColor='DA0000')
            output_data.save(filename="done.xlsx")
            with open('done.xlsx', 'rb') as f:
                file_data = f.read()
            response = HttpResponse(content=file_data, content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment;filename="rept_header.xlsx"'
            return response
        except:
           return redirect('api:index')
